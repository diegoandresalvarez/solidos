# -*- coding: utf-8 -*-

'''
Programa para graficar las líneas de flujo y las equipotenciales
Tenga en cuenta que los bordes de Neumann son líneas de flujo
Y los bordes de Dirichlet

Elaborado por:
* Santiago Beltrán Jaramillo  sbeltran@unal.edu.co 
* Diego Andrés Alvarez Marín  daalvarez@unal.edu.co
'''

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# %%
wb = load_workbook(filename='tablestaca.xlsx', data_only=True, read_only=True)
# ws = wb['Hoja1']
# rango = ws['E3':'BO27']
hoja, coord = list(wb.defined_names['cabezas'].destinations)[0]
rango = wb[hoja][coord]
H = np.array([[celda.value for celda in fila] for fila in rango]).astype(float)
H = np.flipud(H)     # invertir filas de la matriz de arriba a abajo

# %% se definen algunas constantes
K     = 2.5e-08      # [m/s] permeabilidad
delta = 0.5          # [m] espaciado en diferencias finitas

# se definen los ejes
x = np.arange(0, 31 + delta, delta) 
y = np.arange(0, 12 + delta, delta)
xx, yy = np.meshgrid(x, y) # Para el gráfico 2D

# %% Gráfico de la variación de la cabeza hidráulica
plt.figure()
h = plt.contourf(xx, yy, H, 30)                  # colores por niveles
plt.contour(xx, yy, H, h.levels, colors='Black') # curvas de nivel
plt.colorbar(h)                                  # barra de colores
plt.title('Variación de la cabeza hidráulica')
plt.xlabel('Eje x [m]')
plt.ylabel('Eje y [m]')
plt.axis('equal')                                # ejes iguales
plt.tight_layout()                               # figuras ajustadas

# %% Gráfico de las líneas de flujos y las líneas equipotenciales
plt.figure()

# se calcula el campo vectorial de velocidades, usando la ley de Darcy
vy,vx = np.gradient(-K*H, delta, delta)

plt.quiver(xx, yy, vx, vy)                       # campo vectorial de velocidades
plt.contour(xx, yy, H, h.levels, colors='Blue')  # líneas equipotenciales

# se definen los puntos iniciales para las redes de flujo
sx = np.arange(0, 32)
sy = np.full_like(sx, 12)
start = np.c_[sx, sy]

# se grafican las líneas de flujo
plt.streamplot(xx, yy, vx, vy,
                   density=5,
                   color='g',
                   start_points=start,
                   minlength=0.7,
                   maxlength=8)

# se grafica el contorno
cx = [0, 31, 31, 16, 16, 15, 15,  0, 0]
cy = [0,  0, 12, 12,  6,  6, 12, 12, 0]
plt.plot(cx, cy, color='Black', linewidth=3)

plt.title('Red de flujo')
plt.xlabel('Eje x [m]')
plt.ylabel('Eje y [m]')
plt.axis('equal')                                # ejes iguales
plt.tight_layout()                               # figuras ajustadas
plt.show()
