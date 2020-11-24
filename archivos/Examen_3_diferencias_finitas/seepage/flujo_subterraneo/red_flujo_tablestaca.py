# -*- coding: utf-8 -*-
'''
Programa para graficar las líneas de flujo y las equipotenciales
Tenga en cuenta que los bordes de Neumann son líneas de flujo
Y los bordes de Dirichlet
'''

# %%
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# %%
wb = load_workbook(filename='tablestaca.xlsx', data_only=True, read_only=True)
ws = wb['Hoja1']
#rango = wb.defined_names['suelo']
#rango = wb.get_named_range('suelo').destinations
rango = ws['E3':'BO27']
H = np.array([[celda.value for celda in fila] for fila in rango]).astype(float)
H = np.flipud(H) #  % invertir filas de la matriz de arriba a abajo

# %% se definen algunas constantes
K     = 2.5e-08      # [m/s] permeabilidad
delta = 0.5 #Espaciado

# se definen los ejes
x = np.arange(0, 31 + delta, delta) 
y = np.arange(0, 12 + delta, delta)
xx, yy = np.meshgrid(x, y) # Para el gráfico 2D

# %% 
plt.figure()
h = plt.contourf(xx,yy,H, 30) #contourf para el gráfico de contornos

plt.contour(xx, yy, H,
            h.levels, colors='Black') #Para el color de las curvas de nivel

plt.colorbar(h) #Barra de colores
plt.title('Red de Flujo')
plt.xlabel('Eje x')
plt.ylabel('Eje y')
plt.axis('equal') #Para ejes iguales
plt.tight_layout() #Figuras ajustadas

# %%
plt.figure()
vy,vx = np.gradient(-K*H, delta, delta) #Calcula gradiente campo vectorial
plt.quiver(xx, yy, vx, vy) #Grafica la dirección del flujo
plt.contour(xx, yy, H, h.levels, colors='Blue') #Para el color de las curvas

#Definimos inicio líneas de flujo
sx = np.arange(0, 32)
sy = np.full_like(sx, 12)
start = np.c_[sx, sy]

#Graficamos lineas de flujo
plt.streamplot(xx, yy, vx, vy,
                arrowsize=2,
                density=0.7,
                color='g',
                start_points=start,
                minlength=0.7,
                linewidth=2,
                maxlength=8)  #Líneas de flujo (canales)

#Ahora el contorno
cx = [0, 31, 31, 16, 16, 15, 15,  0, 0]
cy = [0,  0, 12, 12,  6,  6, 12, 12, 0]
plt.plot(cx, cy, color='Black', linewidth=3)

plt.axis('equal') #Para ejes iguales
# plt.axis('tight')
plt.axis([-0.9, 32, -0.9, 12.5])
plt.tight_layout() #Figuras ajustadas
plt.show()
